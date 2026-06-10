#!/usr/bin/env python3
"""Rebuild docs picking up ALL sub-question answer columns"""
import openpyxl, subprocess, json, sys

if hasattr(sys.stdout,"reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8",errors="replace")

wb = openpyxl.load_workbook(r"D:\OpenClaw\workspace\temp_interview.xlsx")
MAX_A = 40

def make_doc(sname, skip=0, take=999, label=""):
    ws = wb[sname]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    
    # Build question tree: each Q has a main header and sub-question labels (2,3,4...)
    # Col 5+ defines: Qx (main question) → 2,3,4... (sub-questions) → Qy → 2,3,4...
    q_tree = []  # list of (col, q_text, sub_cols)
    current_q = None
    current_subs = []
    
    for j in range(5, len(hdr)):
        h = hdr[j]
        txt = str(h).strip() if h else ""
        
        # Skip metadata columns
        if any(k in txt.lower() for k in ["开头准备","年龄","职业","城市"]):
            continue
        
        if not txt:
            # Empty header = sub-answer column under current question
            if current_q:
                current_subs.append(j)
            continue
        
        # Check if header is a question (starts with Q/B or is a long sentence)
        is_q = False
        if txt.upper().startswith("Q") and (len(txt) <= 5 or txt[1:].replace(".","").replace(" ","").isdigit()):
            is_q = True
        elif txt.upper().startswith("B") and len(txt) <= 5:
            is_q = True
        elif len(txt) > 8 and not txt.isdigit():
            is_q = True
        
        if is_q:
            if current_q:
                q_tree.append((current_q[0], current_q[1], current_subs))
            current_q = (j, " ".join(txt.split()))
            current_subs = []
        elif current_q and (txt.isdigit() or (len(txt) <= 5 and txt.replace(".","").isdigit())):
            current_subs.append(j)
        else:
            if current_q:
                q_tree.append((current_q[0], current_q[1], current_subs))
                current_q = None
                current_subs = []
    
    # Don't forget the last question
    if current_q:
        q_tree.append((current_q[0], current_q[1], current_subs))
    
    dr = rows[3:]
    group = " ".join(str(rows[1][0]).split()) if len(rows) > 1 and rows[1][0] else ""
    sub = " ".join(str(rows[2][0]).split()) if len(rows) > 2 and rows[2][0] else ""
    
    lines = ["# UE4用户访谈 - {}{}".format(sname, " ({})".format(label) if label else ""), ""]
    if group:
        lines.append("**群组：** " + group)
    if sub:
        lines.append("**子分组：** " + sub)
    lines.append("")
    
    found, added = 0, 0
    for row in dr:
        status = str(row[3]) if len(row) > 3 and row[3] else ""
        if "接受" not in status:
            continue
        found += 1
        if found <= skip:
            continue
        if added >= take:
            break
        added += 1
        
        uid = str(row[1]) if len(row) > 1 and row[1] else "?"
        interviewer = str(row[2]) if len(row) > 2 and row[2] else ""
        
        lines.append("---")
        lines.append("## 受访者 {} (UID: {})".format(added, uid))
        lines.append("")
        if interviewer:
            lines.append("**访谈者：** {}".format(interviewer))
        lines.append("")
        
        qa_count = 0
        for q_col, q_text, sub_cols in q_tree:
            # Main answer
            if q_col < len(row) and row[q_col] and str(row[q_col]).strip():
                a = str(row[q_col]).strip()
                if len(a) > 3:
                    qa_count += 1
                    if len(a) > MAX_A:
                        a = a[:MAX_A] + "..."
                    lines.append("**{}**".format(q_text))
                    lines.append("> {}".format(a.replace("\n", " ")))
                    lines.append("")
            
            # Sub-question answers
            for sc in sub_cols:
                if sc < len(row) and row[sc] and str(row[sc]).strip():
                    a = str(row[sc]).strip()
                    if len(a) > 3:
                        qa_count += 1
                        if len(a) > MAX_A:
                            a = a[:MAX_A] + "..."
                        lines.append("**{} [追问{}]**".format(q_text, sc - q_col))
                        lines.append("> {}".format(a.replace("\n", " ")))
                        lines.append("")
        
        if qa_count == 0:
            lines.append("（无详细问答记录）")
            lines.append("")
    
    if qa_count == 0:
        return None
    
    lines.append("---\n*完整原始数据见 xlsx 表格*")
    return "\n".join(lines)


def create_doc(name, content):
    r = subprocess.run(
        ["dws", "doc", "create", "--name", name, "--markdown", content, "--format", "json"],
        capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=30
    )
    if r.returncode == 0:
        return json.loads(r.stdout).get("docUrl", "?")
    return "FAIL({}b)".format(len(content))


TARGET = 7000

for sname in wb.sheetnames:
    ws = wb[sname]
    rows = list(ws.iter_rows(values_only=True))
    dr = rows[3:]
    users = [r for r in dr if len(r) > 3 and r[3] and "接受" in str(r[3])]
    
    if not users:
        print("{}: 0 users".format(sname))
        continue
    
    # Test content size for 1 user first
    test = make_doc(sname, 0, 1)
    size_per_user = len(test) if test else 10000
    # One user per doc to be safe
    users_per_doc = 1
    
    parts = (len(users) + users_per_doc - 1) // users_per_doc
    print("{}: {} users, ~{}KB/user, {} per doc, {} parts".format(
        sname, len(users), size_per_user // 1000, users_per_doc, parts))
    
    for p in range(parts):
        skip = p * users_per_doc
        label = "P{}".format(p + 1) if parts > 1 else ""
        content = make_doc(sname, skip, users_per_doc, label)
        if not content:
            continue
        name = "【UE4访谈】{}".format(sname)
        if label:
            name += " ({})".format(label)
        url = create_doc(name, content)
        print("  P{}: {}b -> {}".format(p + 1, len(content), url))
    print()
