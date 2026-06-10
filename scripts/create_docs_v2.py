#!/usr/bin/env python3
"""Re-extract interview data correctly with universal column detection"""
import openpyxl, subprocess, json, sys
from datetime import datetime

if hasattr(sys.stdout,"reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8",errors="replace")

wb = openpyxl.load_workbook(r"D:\OpenClaw\workspace\temp_interview.xlsx")
MAX_A = 150

def detect_columns(header):
    """Detect which columns are metadata vs Q&A based on header content."""
    meta_cols = set()
    q_cols = {}
    
    for j, h in enumerate(header):
        if h is None or not str(h).strip():
            continue
        txt = str(h).strip().lower()
        # Identify meta columns by known keywords
        if any(kw in txt for kw in ["排序", "uid", "归属", "联系", "奖励", "开头准备", "年龄", "职业", "城市"]):
            meta_cols.add(j)
        elif txt.startswith("q") and (len(txt) <= 4 or txt[1:].replace(".","").isdigit()):
            q_cols[j] = str(h).strip()
        elif len(txt) > 5 and not txt.isdigit():
            # Long text = question
            q_cols[j] = ' '.join(str(h).strip().split())
    
    return meta_cols, q_cols


def extract_sheet(sname, skip_n=0, max_n=999, label=""):
    ws = wb[sname]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None
    
    header = rows[0]
    q_map = {}
    meta_cols = set()
    
    for j, h in enumerate(header):
        if h is None or not str(h).strip():
            continue
        txt = str(h).strip().lower()
        # Meta columns (not Q&A)
        if any(kw in txt for kw in ["排序", "归属", "联系", "奖励", "开头准备"]):
            meta_cols.add(j)
        # User info columns
        if any(kw in txt for kw in ["年龄", "职业", "城市"]):
            meta_cols.add(j)
        # Q&A columns - anything meta_cols doesn't catch
        if j not in meta_cols and j >= 5:
            q_map[j] = ' '.join(str(h).strip().split())
    
    # If q_map is too small, use a different strategy: all cols after first 5
    if len(q_map) < 3:
        q_map = {}
        for j in range(5, len(header)):
            if header[j] and str(header[j]).strip():
                q_map[j] = ' '.join(str(header[j]).strip().split())[:80]
    
    dr = rows[3:]  # Skip header rows
    group_label = str(rows[1][0]) if len(rows) > 1 and rows[1][0] else ""
    sub_group = str(rows[2][0]) if len(rows) > 2 and rows[2][0] else ""
    
    lines = ["# UE4 - {}{}".format(sname, " ({})".format(label) if label else ""), ""]
    if group_label:
        lines.append("> **{}**".format(group_label))
    if sub_group:
        lines.append("> *{}*".format(sub_group))
    lines.append("")
    
    found, added = 0, 0
    for row in dr:
        status = str(row[3]) if len(row) > 3 and row[3] else ""
        if "接受" not in status:
            continue
        
        found += 1
        if found <= skip_n:
            continue
        if added >= max_n:
            break
        
        added += 1
        uid = str(row[1]) if len(row) > 1 and row[1] else "?"
        interviewer = str(row[2]) if len(row) > 2 and row[2] else ""
        
        lines.append("---")
        lines.append("## 受访者 {} (UID: {})".format(added, uid))
        lines.append("")
        lines.append("**访谈者**：{}".format(interviewer))
        
        # Show available meta info
        meta_parts = []
        for j in range(5, 9):
            if j in meta_cols and j < len(row) and row[j] and str(row[j]).strip():
                meta_parts.append("[{}] {}".format(header[j], str(row[j]).strip()))
        if meta_parts:
            lines.append("  " + " | ".join(meta_parts))
        
        lines.append("")
        
        qa_found = False
        for j in sorted(q_map.keys()):
            if j >= len(row):
                continue
            v = row[j]
            if v is None or not str(v).strip():
                continue
            answer = str(v).strip()
            if len(answer) < 3:
                continue
            qa_found = True
            q_text = q_map[j]
            if len(answer) > MAX_A:
                answer = answer[:MAX_A] + "..."
            lines.append("- **{}**：{}".format(q_text, answer.replace("\n", " ")))
        
        if not qa_found:
            lines.append("（无详细回答）")
        lines.append("")
    
    if added == 0:
        return None
    
    lines.append("---\n*完整数据见原始 xlsx*")
    return "\n".join(lines)


def create_doc(name, content):
    r = subprocess.run(
        ["dws", "doc", "create", "--name", name, "--markdown", content, "--format", "json"],
        capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=30
    )
    if r.returncode == 0:
        return json.loads(r.stdout).get("docUrl", "?")
    return "FAIL({}b)".format(len(content))


TARGET = 7500

for sname in wb.sheetnames:
    ws = wb[sname]
    rows = list(ws.iter_rows(values_only=True))
    dr = rows[3:] if len(rows) > 3 else []
    
    interviewed = [r for r in dr if len(r) > 3 and r[3] and "接受" in str(r[3])]
    if not interviewed:
        print("{}: 0 users".format(sname))
        continue
    
    total_size = sum(sum(len(str(v)) for v in r[9:] if v and len(str(v).strip()) > 3) for r in interviewed)
    print("{}: {} users, ~{}KB".format(sname, len(interviewed), total_size // 1000))
    
    if total_size < TARGET:
        c = extract_sheet(sname)
        if c:
            print("  1 doc: {}b -> {}".format(len(c), create_doc("【UE4】{}".format(sname), c)))
    else:
        # Split by users
        chunk_size = max(1, len(interviewed) // max(1, (total_size // TARGET)))
        for part in range(0, 10):
            skip = part * chunk_size
            if skip >= len(interviewed):
                break
            c = extract_sheet(sname, skip, chunk_size, "Part {}".format(part + 1))
            if c:
                print("  Part {}: {}b -> {}".format(part + 1, len(c), create_doc(
                    "【UE4】{} (Part {})".format(sname, part + 1), c)))
    print()
