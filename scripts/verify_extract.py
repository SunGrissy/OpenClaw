#!/usr/bin/env python3
"""Verify data extraction correctness"""
import openpyxl, sys

if hasattr(sys.stdout,"reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8",errors="replace")

wb = openpyxl.load_workbook(r"D:\OpenClaw\workspace\temp_interview.xlsx")

for sname in wb.sheetnames:
    ws = wb[sname]
    rows = list(ws.iter_rows(values_only=True))
    dr = rows[3:] if len(rows) > 3 else []
    
    interviewed = [r for r in dr if len(r) > 3 and r[3] and "接受" in str(r[3])]
    
    print("{}: total={} rows, interviewed={}".format(sname, len(dr), len(interviewed)))
    
    if interviewed:
        r = interviewed[0]
        print("  Sample: UID={} interviewer={} status={} age={} job={} city={}".format(
            r[1], r[2], r[3], r[6] if len(r)>6 else "", r[7] if len(r)>7 else "", r[8] if len(r)>8 else ""))
        
        # Count Q with answers
        q_count = sum(1 for j in range(9, len(r)) if r[j] and len(str(r[j]).strip()) > 3)
        print("  Q&A columns with content: {}".format(q_count))
    
    # Also count all by status
    statuses = {}
    for r in dr:
        s = str(r[3]) if len(r) > 3 and r[3] else "空"
        statuses[s] = statuses.get(s, 0) + 1
    print("  Statuses: {}".format(dict(sorted(statuses.items()))))
    print()
