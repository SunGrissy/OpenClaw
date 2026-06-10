#!/usr/bin/env python3
"""Rebuild all interview docs with QA on separate lines"""
import openpyxl, subprocess, json, sys

if hasattr(sys.stdout,"reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8",errors="replace")

wb = openpyxl.load_workbook(r"D:\OpenClaw\workspace\temp_interview.xlsx")
MAX_A = 150

def make_doc(sname, skip=0, take=999, label=""):
    ws = wb[sname]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    
    # Identify question columns: cols 5+ where header is not meta
    meta_kw = ["排序","uid","归属","联系","奖励","开头准备"]
    qm = {}  # col -> question text
    for j in range(5, len(hdr)):
        if hdr[j] is None or not str(hdr[j]).strip():
            continue
        txt = str(hdr[j]).strip().lower()
        if any(k in txt for k in meta_kw):
            continue
        # Clean question text
        qtext = " ".join(str(hdr[j]).strip().split())
        qm[j] = qtext
    
    dr = rows[3:]
    group = " ".join(str(rows[1][0]).split()) if len(rows) > 1 and rows[1][0] else ""
    sub = " ".join(str(rows[2][0]).split()) if len(rows) > 2 and rows[2][0] else ""
    
    lines = ["# UE4用户访谈 - {}{}".format(sname, " ({})".format(label) if label else ""), ""]
    if group:
        lines.append("**群组：** " + group)
    if sub:
        lines.append("**子分组：** " + sub)
    lines.append("")
    
    found, added = 0, 0
    for row in dr:
        status = str(row[3]) if len(row) > 3 and row[3] else ""
        if "接受" not in status:
            continue
        found += 1
        if found <= skip:
            continue
        if added >= take:
            break
        added += 1
        
        uid = str(row[1]) if len(row) > 1 and row[1] else "?"
        interviewer = str(row[2]) if len(row) > 2 and row[2] else ""
        
        lines.append("---")
        lines.append("## 受访者 {} (UID: {})".format(added, uid))
        lines.append("")
        if interviewer:
            lines.append("访谈者：{}".format(interviewer))
        lines.append("")
        
        qa_count = 0
        for j in sorted(qm.keys()):
            if j >= len(row):
                continue
            v = row[j]
            if v is None or not str(v).strip():
                continue
            a = str(v).strip()
            if len(a) < 3:
                continue
            qa_count += 1
            qtext = qm[j]
            if len(a) > MAX_A:
                a = a[:MAX_A] + "..."
            # Question as bold header, answer indented on next line
            lines.append("**{}**".format(qtext))
            lines.append("> {}".format(a.replace("\n", " ")))
            lines.append("")
        
        if qa_count == 0:
            lines.append("（无详细问答记录）")
            lines.append("")
    
    lines.append("---\n*完整原始数据见 xlsx 表格*")
    return "\n".join(lines)


def create_and_delete_old(name_base, content, label=""):
    """Create doc with proper content, return URL"""
    full_name = name_base + (" ({})".format(label) if label else "")
    r = subprocess.run(
        ["dws", "doc", "create", "--name", full_name, "--markdown", content, "--format", "json"],
        capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=30
    )
    if r.returncode != 0:
        return "FAIL"
    data = json.loads(r.stdout)
    return data.get("docUrl", "?")


for sname in wb.sheetnames:
    ws = wb[sname]
    rows = list(ws.iter_rows(values_only=True))
    dr = rows[3:]
    users = [r for r in dr if len(r) > 3 and r[3] and "接受" in str(r[3])]
    
    if not users:
        print("{}: 0 users".format(sname))
        continue
    
    # Estimate: ~2KB per user with MAX_A=150, target 6KB per doc
    # So 3 users per doc
    chunk = max(1, len(users) // max(1, (len(users) * 2 // 6)))
    # Actually simpler: fixed 2-3 users per doc
    users_per_doc = 3
    
    parts = (len(users) + users_per_doc - 1) // users_per_doc
    print("{}: {} users, {} parts".format(sname, len(users), parts))
    
    for p in range(parts):
        skip = p * users_per_doc
        label = "Part {}".format(p + 1) if parts > 1 else ""
        content = make_doc(sname, skip, users_per_doc, label)
        url = create_and_delete_old("【UE4访谈】{}".format(sname), content, label)
        print("  Part {}: {}b -> {}".format(p + 1, len(content), url))
    print()
