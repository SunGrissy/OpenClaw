#!/usr/bin/env python3
"""Final rebuild: compact Q&A format with short labels"""
import openpyxl, subprocess, json, sys

if hasattr(sys.stdout,"reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8",errors="replace")

wb = openpyxl.load_workbook(r"D:\OpenClaw\workspace\temp_interview.xlsx")
MAX = 200

def build_qtree(sname):
    ws = wb[sname]
    hdr = list(ws.iter_rows(values_only=True))[0]
    qt = []
    cq, subs = None, []
    qnum = 1
    
    for j in range(5, len(hdr)):
        h = hdr[j]
        txt = str(h).strip() if h else ""
        skip_kw = ["开头准备","年龄","职业","城市"]
        if any(k in txt.lower() for k in skip_kw):
            continue
        if not txt:
            if cq:
                subs.append(j)
            continue
        
        is_q = bool((txt.upper().startswith("Q") and len(txt) <= 5) or (len(txt) > 8 and not txt.isdigit()) or (txt.upper().startswith("B") and len(txt) <= 3))
        
        if is_q:
            if cq:
                qt.append((cq[0], "Q{}".format(qnum), cq[1], subs))
                qnum += 1
            cq = (j, " ".join(txt.split())[:100])
            subs = []
        elif cq and (txt.isdigit() or len(txt) <= 5):
            subs.append(j)
    
    if cq:
        qt.append((cq[0], "Q{}".format(qnum), cq[1], subs))
    return qt

def make(sname, skip=0, take=1, label=""):
    ws = wb[sname]
    rows = list(ws.iter_rows(values_only=True))
    dr = rows[3:]
    qt = build_qtree(sname)
    lines = ["# UE4 - {}{}".format(sname, " ({})".format(label) if label else ""), ""]
    
    found, added = 0, 0
    for row in dr:
        if "接受" not in str(row[3]): continue
        found += 1
        if found <= skip: continue
        if added >= take: break
        added += 1
        uid = str(row[1]) if len(row) > 1 and row[1] else "?"
        iv = str(row[2]) if len(row) > 2 and row[2] else ""
        
        lines.append("---")
        lines.append("## UID: {}".format(uid))
        lines.append("")
        if iv: lines.append("访谈者：" + iv)
        lines.append("")
        
        qc = 0
        for q_col, q_label, q_full, subs in qt:
            ma = ""
            if q_col < len(row) and row[q_col] and str(row[q_col]).strip():
                a = str(row[q_col]).strip()
                if len(a) >= 3:
                    ma = a[:MAX] + ("..." if len(a) > MAX else "")
            
            sa = []
            si = 1
            for sc in subs:
                if sc < len(row) and row[sc] and str(row[sc]).strip():
                    a = str(row[sc]).strip()
                    if len(a) >= 3:
                        sa.append("追问{}：{}".format(si, (a[:MAX] + ("..." if len(a) > MAX else "")).replace("\n"," ")))
                        si += 1
            
            if not ma and not sa: continue
            qc += 1
            lines.append("**{}** {}".format(q_label, q_full[:80]))
            if ma:
                lines.append("    " + ma.replace("\n", " "))
            for sa_line in sa:
                lines.append("    " + sa_line)
        
        if qc == 0:
            lines.append("（无问答）")
            lines.append("")
    
    lines.append("---\n*原始数据见xlsx*")
    return "\n".join(lines)

def create(n, c):
    r = subprocess.run(["dws","doc","create","--name",n,"--markdown",c,"--format","json"],
        capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=30)
    if r.returncode == 0:
        return json.loads(r.stdout).get("docUrl","?")
    return "FAIL({}b)".format(len(c))

TARGET = 8000

for sname in wb.sheetnames:
    test = make(sname, 0, 1)
    size_per = len(test) if test else 8000
    nusers = 0
    ws = wb[sname]
    for row in list(ws.iter_rows(values_only=True))[3:]:
        if "接受" in str(row[3]): nusers += 1
    
    # Determine if we need to split
    users_per_doc = max(1, TARGET // size_per) if size_per else 1
    parts = (nusers + users_per_doc - 1) // users_per_doc
    
    print("{}: {} users, ~{}KB/user, {} per doc, {} parts".format(sname, nusers, size_per//1000, users_per_doc, parts))
    
    for p in range(parts):
        skip = p * users_per_doc
        label = "P{}".format(p+1) if parts > 1 else ""
        c = make(sname, skip, users_per_doc, label)
        if not c: continue
        name = "【UE4】{}".format(sname)
        if label: name += " ({})".format(label)
        url = create(name, c)
        status = "OK" if not url.startswith("FAIL") else "OV"
        print("  {}{}: {}b {}".format(label, "(OVER)" if url.startswith("FAIL") else "", len(c), url[:50]))
    print()
